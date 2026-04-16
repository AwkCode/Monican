#!/usr/bin/env python3
"""
Monican — Outreach Drafter
Reads prospects.xlsx and generates personalized cold emails + LinkedIn messages.
Usage: python3 agents/outreach_drafter.py [--priority HIGH|MEDIUM|LOW] [--limit N]
"""

import openpyxl
import os
import sys
from datetime import datetime

PROSPECTS_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "prospects.xlsx")
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "outreach_drafts")


def load_prospects(priority_filter=None):
    wb = openpyxl.load_workbook(PROSPECTS_FILE)
    ws = wb["Prospects"]

    headers = [cell.value for cell in ws[1]]
    prospects = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        prospect = dict(zip(headers, row))
        if prospect.get("Agency Name") is None:
            continue
        if priority_filter and prospect.get("Priority", "").upper() != priority_filter.upper():
            continue
        # Skip already contacted
        if prospect.get("Outreach Status") and prospect["Outreach Status"].strip():
            continue
        prospects.append(prospect)

    return prospects


def get_first_name(full_name):
    if not full_name:
        return ""
    return full_name.strip().split()[0]


def draft_email(prospect):
    name = prospect.get("Broker/Owner", "")
    first_name = get_first_name(name)
    agency = prospect.get("Agency Name", "")
    town = prospect.get("Town", "")
    agents = prospect.get("Est. Agents", "")
    details = prospect.get("Notable Details", "")
    distance = prospect.get("Miles from Bolton", "")

    # Choose template based on context
    if distance and str(distance).strip() in ["0", "5", "6"]:
        # Local angle for very nearby agencies
        template = "local"
    elif agents and any(x in str(agents) for x in ["10", "15", "20", "30", "40", "50"]):
        # Specific pain point for larger teams
        template = "pain_point"
    else:
        # Default free audit offer
        template = "audit"

    output = f"## {agency} — {town}\n"
    output += f"**Contact:** {name}\n"
    output += f"**Template:** {template}\n\n"

    if template == "local":
        output += f"**Subject:** Fellow Bolton local — quick AI question\n\n"
        output += f"Hi {first_name},\n\n"
        output += f"I'm based right here in Bolton and I've been building AI automation tools for small businesses. "
        output += f"Real estate is one of the industries where I keep seeing the biggest opportunities — especially around lead follow-up, listing descriptions, and transaction coordination.\n\n"
        output += f"I'd love to buy you a coffee and hear about how {agency} handles the admin side of things. "
        output += f"Not a sales pitch — genuinely curious about where AI could help local agencies like yours.\n\n"
        output += f"Free anytime this week?\n\n"
        output += f"Daniel Weadock\n"

    elif template == "pain_point":
        output += f"**Subject:** Saving your agents 10+ hours/week on admin\n\n"
        output += f"Hi {first_name},\n\n"
        output += f"I noticed {agency} has about {agents} agents — which means your team is probably spending a huge chunk of time on lead responses, listing writeups, and chasing documents for transactions.\n\n"
        output += f"I build AI systems that handle that automatically. One team I've worked with cut their lead response time from hours to minutes and stopped writing listing descriptions entirely — the AI drafts them from property details in seconds.\n\n"
        output += f"I'd love to show you what this looks like for {agency}. Would a 20-minute call work sometime this week?\n\n"
        output += f"Daniel Weadock\nMonican\n"

    else:
        output += f"**Subject:** Quick question about {agency}'s admin workflow\n\n"
        output += f"Hi {first_name},\n\n"
        output += f"I help real estate teams automate their most time-consuming admin work using AI — things like lead follow-up, listing descriptions, transaction tracking, and inbox management.\n\n"
        output += f"I'm offering free workflow audits to a handful of agencies in the {town} area this month. "
        output += f"It's a 30-minute conversation where I map out where your team is spending the most time on repetitive tasks and show you exactly what could be automated.\n\n"
        output += f"No pitch, no commitment — just a clear picture of where AI could save your agents real hours every week.\n\n"
        output += f"Would you be open to a quick call this week or next?\n\n"
        output += f"Best,\nDaniel Weadock\nMonican\n"

    # Add personalization notes
    if details:
        output += f"\n**Personalization notes:** {details}\n"

    return output


def draft_linkedin(prospect):
    name = prospect.get("Broker/Owner", "")
    first_name = get_first_name(name)
    agency = prospect.get("Agency Name", "")
    town = prospect.get("Town", "")
    distance = prospect.get("Miles from Bolton", "")

    output = f"\n### LinkedIn Connection Request\n"

    if distance and str(distance).strip() in ["0", "5", "6", "8"]:
        output += f"Hi {first_name} — I'm based in Bolton and I help real estate teams automate admin work with AI. "
        output += f"I'd love to connect and learn about how {agency} handles the ops side. No pitch, just curious!\n"
    else:
        output += f"Hi {first_name} — I build AI tools that save real estate agents 10+ hrs/week on lead follow-up, listings, and transaction admin. "
        output += f"Would love to connect and share what I'm seeing in the space.\n"

    return output


def main():
    priority_filter = None
    limit = None

    args = sys.argv[1:]
    for i, arg in enumerate(args):
        if arg == "--priority" and i + 1 < len(args):
            priority_filter = args[i + 1]
        if arg == "--limit" and i + 1 < len(args):
            limit = int(args[i + 1])

    prospects = load_prospects(priority_filter)
    if limit:
        prospects = prospects[:limit]

    if not prospects:
        print("No prospects found matching criteria.")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"outreach_{date_str}"
    if priority_filter:
        filename += f"_{priority_filter.lower()}"
    filename += ".md"

    filepath = os.path.join(OUTPUT_DIR, filename)

    output = f"# Outreach Drafts — {date_str}\n"
    output += f"**Filter:** {priority_filter or 'All'} priority\n"
    output += f"**Count:** {len(prospects)} prospects\n\n"
    output += "---\n\n"

    for i, prospect in enumerate(prospects, 1):
        output += f"# Prospect {i}: {prospect.get('Agency Name', 'Unknown')}\n\n"
        output += draft_email(prospect)
        output += draft_linkedin(prospect)
        output += "\n\n---\n\n"

    with open(filepath, "w") as f:
        f.write(output)

    print(f"Generated {len(prospects)} outreach drafts → {filepath}")


if __name__ == "__main__":
    main()
